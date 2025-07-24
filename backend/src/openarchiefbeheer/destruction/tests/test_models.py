import os
from datetime import date, datetime
from unittest.mock import patch

from django.contrib.auth.models import Group
from django.core.exceptions import ObjectDoesNotExist
from django.core.files.base import ContentFile
from django.test import TestCase
from django.utils import timezone
from django.utils.translation import gettext

from freezegun import freeze_time
from openpyxl import load_workbook
from privates.test import temp_private_root
from requests import HTTPError
from requests_mock import Mocker
from zgw_consumers.constants import APITypes
from zgw_consumers.test.factories import ServiceFactory

from openarchiefbeheer.config.models import ArchiveConfig
from openarchiefbeheer.logging import logevent
from openarchiefbeheer.utils.results_store import ResultStore
from openarchiefbeheer.utils.tests.mixins import ClearCacheMixin
from openarchiefbeheer.zaken.models import Zaak
from openarchiefbeheer.zaken.tests.factories import ZaakFactory

from ...accounts.tests.factories import UserFactory
from ..constants import (
    InternalStatus,
    ListItemStatus,
    ListStatus,
    ReviewDecisionChoices,
)
from .factories import (
    DestructionListCoReviewFactory,
    DestructionListFactory,
    DestructionListItemFactory,
    DestructionListReviewFactory,
    ReviewResponseFactory,
)


class DestructionListItemTest(ClearCacheMixin, TestCase):
    def test_set_status(self):
        destruction_list = DestructionListFactory.create()

        with freeze_time("2024-05-02T16:00:00+02:00"):
            destruction_list.set_status(ListItemStatus.removed)

        destruction_list.refresh_from_db()

        self.assertEqual(
            destruction_list.status_changed,
            timezone.make_aware(datetime(2024, 5, 2, 16, 0)),
        )

    def test_process_deletion_zaak_not_found(self):
        ServiceFactory.create(
            api_root="http://zaken.nl/api/v1/", label="Open Zaak - Zaken API"
        )
        item = DestructionListItemFactory.create(
            with_zaak=True,
            zaak__url="http://zaken.nl/api/v1/zaken/111-111-111",
        )

        item.process_deletion()

        item.refresh_from_db()

        self.assertEqual(item.processing_status, InternalStatus.failed)

    def test_process_deletion(self):
        ServiceFactory.create(
            api_root="http://zaken.nl/api/v1/", label="Open Zaak - Zaken API"
        )
        item = DestructionListItemFactory.create(
            with_zaak=True,
            zaak__url="http://zaken.nl/api/v1/zaken/111-111-111",
            zaak__omschrijving="Test description",
        )

        with patch(
            "openarchiefbeheer.destruction.models.delete_zaak_and_related_objects"
        ) as m_delete_in_openzaak:
            item.process_deletion()

        m_delete_in_openzaak.assert_called_once()
        kwargs = m_delete_in_openzaak.call_args_list[0].kwargs
        self.assertEqual(kwargs["zaak"].url, "http://zaken.nl/api/v1/zaken/111-111-111")
        self.assertEqual(kwargs["result_store"].store.pk, item.pk)

        item.refresh_from_db()

        self.assertEqual(item.processing_status, InternalStatus.succeeded)

        with self.assertRaises(ObjectDoesNotExist):
            Zaak.objects.get(url="http://zaken.nl/api/v1/zaken/111-111-111")

    def test_keeping_zaak_url_in_sync(self):
        zaak = ZaakFactory.create(url="http://zaken.nl/1")
        item = DestructionListItemFactory.create(
            with_zaak=True, zaak__url="http://zaken.nl/2"
        )

        self.assertEqual(item._zaak_url, "http://zaken.nl/2")

        item.zaak = zaak
        item.save()

        item.refresh_from_db()

        self.assertEqual(item._zaak_url, "http://zaken.nl/1")

    def test_zaak_data_present_after_deletion(self):
        ServiceFactory.create(
            api_root="http://zaken.nl/api/v1/", label="Open Zaak - Zaken API"
        )
        item = DestructionListItemFactory.create(
            with_zaak=True,
            zaak__url="http://zaken.nl/api/v1/zaken/111-111-111",
            zaak__omschrijving="Test description",
            zaak__zaaktype="http://catalogi.nl/api/v1/zaaktypen/111-111-111",
            zaak__identificatie="ZAAK-01",
            zaak__startdatum=date(2020, 1, 1),
            zaak__einddatum=date(2022, 1, 1),
            zaak__resultaat="http://zaken.nl/api/v1/resultaten/111-111-111",
        )

        with patch(
            "openarchiefbeheer.destruction.models.delete_zaak_and_related_objects"
        ):
            item.process_deletion()

        item.refresh_from_db()

        self.assertEqual(item.processing_status, InternalStatus.succeeded)
        self.assertEqual(
            item.extra_zaak_data["url"], "http://zaken.nl/api/v1/zaken/111-111-111"
        )
        self.assertEqual(
            item.extra_zaak_data["bronapplicatie"], "Open Zaak - Zaken API"
        )
        self.assertEqual(item.extra_zaak_data["omschrijving"], "Test description")
        self.assertEqual(item.extra_zaak_data["identificatie"], "ZAAK-01")
        self.assertEqual(item.extra_zaak_data["startdatum"], "2020-01-01")
        self.assertEqual(item.extra_zaak_data["einddatum"], "2022-01-01")

        self.assertEqual(item.extra_zaak_data["zaaktype"]["uuid"], "111-111-111")
        self.assertEqual(
            item.extra_zaak_data["zaaktype"]["url"],
            "http://catalogi.nl/api/v1/zaaktypen/111-111-111",
        )
        self.assertEqual(
            item.extra_zaak_data["zaaktype"]["identificatie"], "ZAAKTYPE-01"
        )
        self.assertEqual(
            item.extra_zaak_data["zaaktype"]["selectielijst_procestype"]["naam"],
            "Evaluatie uitvoeren",
        )

        self.assertEqual(
            item.extra_zaak_data["resultaat"],
            {
                "url": "http://zaken-api.nl/zaken/api/v1/resultaten/111-111-111",
                "resultaattype": {"omschrijving": "This is a result type"},
            },
        )


class ReviewResponseTests(TestCase):
    def test_derive_status(self):
        review_response = ReviewResponseFactory.create()

        self.assertEqual(
            review_response._derive_status(
                [InternalStatus.new, InternalStatus.new, InternalStatus.new]
            ),
            InternalStatus.new,
        )
        self.assertEqual(
            review_response._derive_status(
                [
                    InternalStatus.succeeded,
                    InternalStatus.succeeded,
                    InternalStatus.succeeded,
                ]
            ),
            InternalStatus.succeeded,
        )
        self.assertEqual(
            review_response._derive_status(
                [
                    InternalStatus.failed,
                    InternalStatus.succeeded,
                    InternalStatus.succeeded,
                ]
            ),
            InternalStatus.failed,
        )
        self.assertEqual(
            review_response._derive_status(
                [
                    InternalStatus.processing,
                    InternalStatus.succeeded,
                    InternalStatus.succeeded,
                ]
            ),
            InternalStatus.processing,
        )
        self.assertEqual(
            review_response._derive_status(
                [
                    InternalStatus.queued,
                    InternalStatus.succeeded,
                    InternalStatus.succeeded,
                ]
            ),
            InternalStatus.queued,
        )
        self.assertEqual(
            review_response._derive_status(
                [
                    InternalStatus.queued,
                    InternalStatus.failed,
                    InternalStatus.processing,
                ]
            ),
            InternalStatus.failed,
        )
        self.assertEqual(
            review_response._derive_status(
                [InternalStatus.queued, InternalStatus.new, InternalStatus.processing]
            ),
            InternalStatus.processing,
        )
        self.assertEqual(
            review_response._derive_status(
                [InternalStatus.new, InternalStatus.new, InternalStatus.succeeded]
            ),
            InternalStatus.processing,
        )


TEST_DATA_ARCHIVE_CONFIG = {
    "bronorganisatie": "000000000",
    "zaaktype": "ZAAKTYPE-2018-0000000002",
    "statustype": "http://localhost:8003/catalogi/api/v1/statustypen/835a2a13-f52f-4339-83e5-b7250e5ad016",
    "resultaattype": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
    "informatieobjecttype": "http://localhost:8003/catalogi/api/v1/informatieobjecttypen/9dee6712-122e-464a-99a3-c16692de5485",
}

TEST_DATA_ARCHIVE_CONFIG_PARTIAL = {
    "bronorganisatie": "000000000",
    "zaaktype": "ZAAKTYPE-2018-0000000002",
    "informatieobjecttype": "http://localhost:8003/catalogi/api/v1/informatieobjecttypen/9dee6712-122e-464a-99a3-c16692de5485",
    "resultaattype": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
}


@temp_private_root()
class DestructionListTest(TestCase):
    def test_has_long_review_process(self):
        destruction_list = DestructionListFactory.create()

        DestructionListItemFactory.create(
            destruction_list=destruction_list,
            with_zaak=True,
            zaak__zaaktype="http://catalogi-api.nl/zaaktypen/111-111-111",
            zaak__post___expand={
                "zaaktype": {
                    "identificatie": "ZAAKTYPE-01",
                    "omschrijving": "ZAAKTYPE-01",
                    "versiedatum": "2024-01-01",
                    "url": "http://catalogue-api.nl/zaaktypen/111-111-111",
                }
            },
        )
        DestructionListItemFactory.create(
            destruction_list=destruction_list,
            with_zaak=True,
            zaak__zaaktype="http://catalogi-api.nl/zaaktypen/111-111-111",
            zaak__post___expand={
                "zaaktype": {
                    "identificatie": "ZAAKTYPE-01",
                    "omschrijving": "ZAAKTYPE-01",
                    "versiedatum": "2024-01-01",
                    "url": "http://catalogue-api.nl/zaaktypen/111-111-111",
                }
            },
        )
        DestructionListItemFactory.create(
            destruction_list=destruction_list,
            with_zaak=True,
            zaak__zaaktype="http://catalogi-api.nl/zaaktypen/222-222-222",
            zaak__post___expand={
                "zaaktype": {
                    "identificatie": "ZAAKTYPE-02",
                    "omschrijving": "ZAAKTYPE-02",
                    "versiedatum": "2024-01-02",
                    "url": "http://catalogue-api.nl/zaaktypen/222-222-222",
                }
            },
        )

        with patch(
            "openarchiefbeheer.destruction.models.ArchiveConfig.get_solo",
            return_value=ArchiveConfig(zaaktypes_short_process=["ZAAKTYPE-01"]),
        ):
            has_short_review_process = destruction_list.has_short_review_process()

        self.assertFalse(has_short_review_process)

    def test_has_short_review_process(self):
        destruction_list = DestructionListFactory.create()

        DestructionListItemFactory.create(
            destruction_list=destruction_list,
            with_zaak=True,
            zaak__zaaktype="http://catalogi-api.nl/zaaktypen/1",
            zaak__post___expand={
                "zaaktype": {
                    "identificatie": "ZAAKTYPE-01",
                    "omschrijving": "ZAAKTYPE-01",
                    "versiedatum": "2024-01-01",
                    "url": "http://catalogue-api.nl/zaaktypen/111-111-111",
                }
            },
        )
        DestructionListItemFactory.create(
            destruction_list=destruction_list,
            with_zaak=True,
            zaak__zaaktype="http://catalogi-api.nl/zaaktypen/1",
            zaak__post___expand={
                "zaaktype": {
                    "identificatie": "ZAAKTYPE-01",
                    "omschrijving": "ZAAKTYPE-01",
                    "versiedatum": "2024-01-01",
                    "url": "http://catalogue-api.nl/zaaktypen/111-111-111",
                }
            },
        )
        DestructionListItemFactory.create(
            destruction_list=destruction_list,
            with_zaak=True,
            zaak__zaaktype="http://catalogi-api.nl/zaaktypen/222-222-222",
            zaak__post___expand={
                "zaaktype": {
                    "identificatie": "ZAAKTYPE-02",
                    "omschrijving": "ZAAKTYPE-02",
                    "versiedatum": "2024-01-02",
                    "url": "http://catalogue-api.nl/zaaktypen/222-222-222",
                }
            },
        )

        with patch(
            "openarchiefbeheer.destruction.models.ArchiveConfig.get_solo",
            return_value=ArchiveConfig(
                zaaktypes_short_process=[
                    "ZAAKTYPE-01",
                    "ZAAKTYPE-02",
                ]
            ),
        ):
            has_short_review_process = destruction_list.has_short_review_process()

        self.assertTrue(has_short_review_process)

    def test_generate_destruction_report(self):
        record_manager = UserFactory.create(
            first_name="John",
            last_name="Doe",
            username="jdoe1",
            post__can_start_destruction=True,
        )
        destruction_list = DestructionListFactory.create(
            status=ListStatus.deleted,
            end=datetime(2024, 12, 2, 12, tzinfo=timezone.get_default_timezone()),
        )
        with freeze_time("2024-12-01T12:00:00+01:00"):
            logevent.destruction_list_deletion_triggered(
                destruction_list, record_manager
            )
        DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            status=ListItemStatus.suggested,
            destruction_list=destruction_list,
            extra_zaak_data={
                "bronapplicatie": "Open Zaak",
                "url": "http://zaken.nl/api/v1/zaken/111-111-111",
                "omschrijving": "Test description 1",
                "identificatie": "ZAAK-01",
                "startdatum": "2020-01-01",
                "einddatum": "2022-01-01",
                "resultaat": {
                    "url": "http://zaken.nl/api/v1/resultaten/111-111-111",
                    "resultaattype": {
                        "url": "http://catalogue-api.nl/catalogi/api/v1/resultaattypen/111-111-111",
                        "archiefactietermijn": "P1D",
                        "omschrijving": "Resulttype 0",
                    },
                },
                "zaaktype": {
                    "uuid": "111-111-111",
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-111",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "naam": "Plannen opstellen",
                        "nummer": 1.1,
                        "jaar": 2020,
                    },
                },
                "selectielijstklasse": "3.2 - Niet vastgesteld - vernietigen",
                "selectielijstklasse_versie": "2020",
            },
        )
        DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            status=ListItemStatus.suggested,
            destruction_list=destruction_list,
            extra_zaak_data={
                "bronapplicatie": "Open Zaak",
                "url": "http://zaken.nl/api/v1/zaken/111-111-222",
                "omschrijving": "Test description 2",
                "identificatie": "ZAAK-02",
                "startdatum": "2020-01-02",
                "einddatum": "2022-01-02",
                "resultaat": {
                    "url": "http://zaken.nl/api/v1/resultaten/111-111-222",
                    "resultaattype": {
                        "url": "http://catalogue-api.nl/catalogi/api/v1/resultaattypen/111-111-111",
                        "archiefactietermijn": "P1D",
                        "omschrijving": "Resulttype 0",
                    },
                },
                "zaaktype": {
                    "uuid": "111-111-111",
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-111",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "nummer": 1,
                        "naam": "Beleid en regelgeving opstellen",
                        "jaar": 2024,
                    },
                },
                "selectielijstklasse": "3.2 - Niet vastgesteld - vernietigen",
                "selectielijstklasse_versie": "2024",
            },
        )
        DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            destruction_list=destruction_list,
            extra_zaak_data={
                "bronapplicatie": "Open Zaak",
                "url": "http://zaken.nl/api/v1/zaken/111-111-333",
                "omschrijving": "Test description 3",
                "identificatie": "ZAAK-03",
                "startdatum": "2020-01-03",
                "einddatum": "2022-01-03",
                "resultaat": None,
                "zaaktype": {
                    "uuid": "111-111-222",
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-222",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "naam": "Instellen en inrichten organisatie",
                        "nummer": 1.0,
                        "jaar": 2022,
                    },
                },
                "selectielijstklasse": "1.1 - Ingericht - vernietigen",
                "selectielijstklasse_versie": "2022",
            },
        )

        destruction_list.generate_destruction_report()

        destruction_list.refresh_from_db()

        wb = load_workbook(filename=destruction_list.destruction_report.path)
        sheet_deleted_zaken = wb[gettext("Deleted zaken")]
        rows = list(sheet_deleted_zaken.iter_rows(values_only=True))

        self.assertEqual(len(rows), 4)
        self.assertEqual(
            rows[0],
            (
                "Bronapplicatie",
                "Zaaktype Omschrijving",
                "Zaaktype UUID",
                "Zaaktype Identificatie",
                "Resultaat",
                "Zaak Identificatie",
                "Zaak Startdatum",
                "Zaak Einddatum",
                "Selectielijstklasse",
                "Selectielijst versie",
            ),
        )

        # We don't know the order in which the zaken are retrieved in the db (they are not sorted).
        data_rows = sorted(rows[1:], key=lambda row: row[0])

        self.assertEqual(
            data_rows[0],
            (
                "Open Zaak",
                "Tralala zaaktype",
                "111-111-111",
                None,  # pyopenxl reads empty values as None
                "Resulttype 0",
                "ZAAK-01",
                "2020-01-01",
                "2022-01-01",
                "3.2 - Niet vastgesteld - vernietigen",
                "2020",
            ),
        )
        self.assertEqual(
            data_rows[1],
            (
                "Open Zaak",
                "Tralala zaaktype",
                "111-111-111",
                None,
                "Resulttype 0",
                "ZAAK-02",
                "2020-01-02",
                "2022-01-02",
                "3.2 - Niet vastgesteld - vernietigen",
                "2024",
            ),
        )
        self.assertEqual(
            data_rows[2],
            (
                "Open Zaak",
                "Tralala zaaktype",
                "111-111-222",
                None,
                None,
                "ZAAK-03",
                "2020-01-03",
                "2022-01-03",
                "1.1 - Ingericht - vernietigen",
                "2022",
            ),
        )

        sheet_process_details = wb[gettext("Process details")]
        rows = list(sheet_process_details.iter_rows(values_only=True))

        self.assertEqual(
            rows[0][:5],
            (
                gettext("Date/Time starting destruction"),
                gettext("Date/Time of destruction"),
                gettext("User who started the destruction"),
                gettext("Groups"),
                gettext("Number of deleted cases"),
            ),
        )
        self.assertEqual(
            rows[1][:5],
            (
                "2024-12-01 12:00+01:00",
                "2024-12-02 12:00+01:00",
                "John Doe (jdoe1)",
                None,
                3,
            ),
        )

    def test_generate_destruction_report_with_cases_excluded_from_list(self):
        author = UserFactory.create(post__can_start_destruction=True)
        destruction_list = DestructionListFactory.create(
            author=author,
            status=ListStatus.deleted,
            end=datetime(2024, 12, 2, 12, tzinfo=timezone.get_default_timezone()),
        )
        DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            status=ListItemStatus.suggested,
            destruction_list=destruction_list,
            extra_zaak_data={
                "bronorganisatie": "Openzaak",
                "url": "http://zaken.nl/api/v1/zaken/111-111-111",
                "omschrijving": "Test description 3",
                "identificatie": "ZAAK-01",
                "startdatum": "2020-01-03",
                "einddatum": "2022-01-03",
                "resultaat": None,
                "zaaktype": {
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-222",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "naam": "Instellen en inrichten organisatie",
                        "nummer": 1,
                        "jaar": 2000,
                    },
                },
            },
        )
        DestructionListItemFactory.create(
            processing_status=InternalStatus.new,
            status=ListItemStatus.removed,  # Case no longer in destruction list
            destruction_list=destruction_list,
            extra_zaak_data={
                "bronorganisatie": "Openzaak",
                "url": "http://zaken.nl/api/v1/zaken/222-222-222",
                "omschrijving": "Test description 3",
                "identificatie": "ZAAK-02",
                "startdatum": "2020-01-03",
                "einddatum": "2022-01-03",
                "resultaat": None,
                "zaaktype": {
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-222",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "naam": "Instellen en inrichten organisatie",
                        "nummer": 1.2,
                        "jaar": 2004,
                    },
                },
            },
        )
        DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            destruction_list=destruction_list,
            status=ListItemStatus.suggested,
            extra_zaak_data={
                "url": "http://zaken.nl/api/v1/zaken/333-333-333",
                "omschrijving": "Test description 3",
                "identificatie": "ZAAK-03",
                "startdatum": "2020-01-03",
                "einddatum": "2022-01-03",
                "resultaat": None,
                "zaaktype": {
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-222",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "naam": "Instellen en inrichten organisatie",
                        "nummer": 1.2,
                        "jaar": 2004,
                    },
                },
            },
        )
        logevent.destruction_list_deletion_triggered(destruction_list, author)

        destruction_list.generate_destruction_report()

        destruction_list.refresh_from_db()

        wb = load_workbook(filename=destruction_list.destruction_report.path)
        sheet_deleted_zaken = wb[gettext("Deleted zaken")]
        rows = list(sheet_deleted_zaken.iter_rows(values_only=True))

        self.assertEqual(len(rows), 3)
        self.assertEqual(
            rows[0],
            (
                "Bronapplicatie",
                "Zaaktype Omschrijving",
                "Zaaktype UUID",
                "Zaaktype Identificatie",
                "Resultaat",
                "Zaak Identificatie",
                "Zaak Startdatum",
                "Zaak Einddatum",
                "Selectielijstklasse",
                "Selectielijst versie",
            ),
        )
        deleted_zaken_ids = sorted([rows[1][5], rows[2][5]])
        self.assertEqual(
            deleted_zaken_ids,
            ["ZAAK-01", "ZAAK-03"],
        )

    def test_generate_destruction_report_with_multiple_logs(self):
        record_manager1 = UserFactory.create(
            first_name="John",
            last_name="Doe",
            username="jdoe1",
            post__can_start_destruction=True,
        )
        record_manager2 = UserFactory.create(
            first_name="Jane",
            last_name="Doe",
            username="jdoe2",
            post__can_start_destruction=True,
        )
        destruction_list = DestructionListFactory.create(
            status=ListStatus.deleted,
            end=datetime(2024, 12, 2, 12, tzinfo=timezone.get_default_timezone()),
        )
        with freeze_time("2024-12-01T12:00:00+01:00"):
            logevent.destruction_list_deletion_triggered(
                destruction_list, record_manager1
            )
        with freeze_time("2024-12-01T12:30:00+01:00"):
            logevent.destruction_list_deletion_triggered(
                destruction_list, record_manager2
            )

        destruction_list.generate_destruction_report()

        destruction_list.refresh_from_db()

        wb = load_workbook(filename=destruction_list.destruction_report.path)
        sheet_deleted_zaken = wb[gettext("Process details")]
        rows = list(sheet_deleted_zaken.iter_rows(values_only=True))

        self.assertEqual(len(rows), 4)  # No reviews logs
        self.assertEqual(
            rows[0][:5],
            (
                gettext("Date/Time starting destruction"),
                gettext("Date/Time of destruction"),
                gettext("User who started the destruction"),
                gettext("Groups"),
                gettext("Number of deleted cases"),
            ),
        )
        self.assertEqual(
            rows[1][:5],
            (
                "2024-12-01 12:30+01:00",
                "2024-12-02 12:00+01:00",
                "Jane Doe (jdoe2)",
                None,
                0,
            ),
        )

    def test_generate_destruction_report_review_process(self):
        author = UserFactory.create(post__can_start_destruction=True)
        reviewer = UserFactory.create(
            first_name="John",
            last_name="Doe",
            username="jdoe1",
            post__can_review_destruction=True,
        )
        reviewer_group, created = Group.objects.get_or_create(name="Reviewer")
        reviewer.groups.add(reviewer_group)
        archivist = UserFactory.create(
            first_name="Alice",
            last_name="Wonderland",
            username="awonderland1",
            post__can_review_final_list=True,
        )
        archivist_group, created = Group.objects.get_or_create(name="Archivist")
        archivist.groups.add(archivist_group)

        destruction_list = DestructionListFactory.create(
            status=ListStatus.deleted,
            author=author,
            end=datetime(2024, 12, 2, 12, tzinfo=timezone.get_default_timezone()),
        )

        review_reviewer_rejected = DestructionListReviewFactory.create(
            author=reviewer, decision=ReviewDecisionChoices.rejected
        )
        review_reviewer_accepted = DestructionListReviewFactory.create(
            author=reviewer, decision=ReviewDecisionChoices.accepted
        )
        review_archivist_accepted = DestructionListReviewFactory.create(
            author=archivist, decision=ReviewDecisionChoices.accepted
        )

        logevent.destruction_list_created(
            destruction_list, author, reviewer
        )  # A log with a different template (should NOT be present in the report)

        logevent.destruction_list_reviewed(
            destruction_list,
            review_reviewer_rejected,
            comment="Terrible",
            user=reviewer,
        )  # A rejection (should NOT be present in the report)
        logevent.destruction_list_deletion_triggered(destruction_list, author)
        with freeze_time("2024-05-02T16:00:00+02:00"):
            logevent.destruction_list_reviewed(
                destruction_list,
                review_reviewer_accepted,
                comment="Nice",
                user=reviewer,
            )
        with freeze_time("2024-05-08T09:45:00+02:00"):
            logevent.destruction_list_reviewed(
                destruction_list,
                review_archivist_accepted,
                comment="Ship it",
                user=archivist,
            )

        destruction_list.generate_destruction_report()

        wb = load_workbook(filename=destruction_list.destruction_report.path)
        sheet_deleted_zaken = wb[gettext("Process details")]
        rows = list(sheet_deleted_zaken.iter_rows(values_only=True))

        self.assertEqual(len(rows), 6)

        self.assertEqual(
            rows[4][:4],
            (
                "Reviewer",
                "John Doe (jdoe1)",
                "2024-05-02 16:00+02:00",
                gettext("Has approved"),
            ),
        )
        self.assertEqual(
            rows[5][:4],
            (
                "Archivist",
                "Alice Wonderland (awonderland1)",
                "2024-05-08 09:45+02:00",
                gettext("Has approved"),
            ),
        )

    def test_zaak_creation_skipped_if_internal_status_succeeded(self):
        destruction_list = DestructionListFactory.create(
            processing_status=InternalStatus.succeeded
        )

        with patch(
            "openarchiefbeheer.destruction.utils.create_zaak_for_report"
        ) as m_zaak:
            destruction_list.create_report_zaak()

        m_zaak.assert_not_called()

    @Mocker()
    def test_failure_during_zaak_creation(self, m):
        destruction_list = DestructionListFactory.create()
        ServiceFactory.create(
            api_type=APITypes.zrc,
            api_root="http://localhost:8003/zaken/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.ztc,
            api_root="http://localhost:8003/catalogi/api/v1",
        )

        m.get(
            "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
            json={
                "url": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
                "selectielijstklasse": "https://selectielijst.openzaak.nl/api/v1/resultaten/e939c1ad-32e4-409b-a716-6d7d6e7df892",
            },
        )
        m.get(
            "http://localhost:8003/catalogi/api/v1/zaaktypen?identificatie=ZAAKTYPE-2018-0000000002",
            json={
                "count": 1,
                "results": [
                    {
                        "url": "http://localhost:8003/catalogi/api/v1/zaaktypen/ce9feadd-00cb-46c8-a0ef-1d1dfc78586a",
                        "identificatie": "ZAAKTYPE-2018-0000000002",
                        "omschrijving": "Destruction confirmation type",
                        "beginGeldigheid": "2025-03-21",
                    },
                ],
            },
        )
        m.post("http://localhost:8003/zaken/api/v1/zaken", status_code=500)

        with (
            patch(
                "openarchiefbeheer.destruction.utils.ArchiveConfig.get_solo",
                return_value=ArchiveConfig(**TEST_DATA_ARCHIVE_CONFIG),
            ),
            self.assertRaises(HTTPError),
        ):
            destruction_list.create_report_zaak()

        destruction_list.refresh_from_db()

        self.assertEqual(destruction_list.zaak_destruction_report_url, "")

    @Mocker()
    def test_no_statustype_configured(self, m):
        destruction_list = DestructionListFactory.create()
        ServiceFactory.create(
            api_type=APITypes.zrc,
            api_root="http://localhost:8003/zaken/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.ztc,
            api_root="http://localhost:8003/catalogi/api/v1",
        )

        m.get(
            "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
            json={
                "url": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
                "selectielijstklasse": "https://selectielijst.openzaak.nl/api/v1/resultaten/e939c1ad-32e4-409b-a716-6d7d6e7df892",
            },
        )
        m.get(
            "http://localhost:8003/catalogi/api/v1/zaaktypen?identificatie=ZAAKTYPE-2018-0000000002",
            json={
                "count": 1,
                "results": [
                    {
                        "url": "http://localhost:8003/catalogi/api/v1/zaaktypen/ce9feadd-00cb-46c8-a0ef-1d1dfc78586a",
                        "identificatie": "ZAAKTYPE-2018-0000000002",
                        "omschrijving": "Destruction confirmation type",
                        "beginGeldigheid": "2025-03-21",
                    },
                ],
            },
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/zaken",
            json={"url": "http://localhost:8003/zaken/api/v1/zaken/111-111-111"},
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/resultaten",
            json={"url": "http://localhost:8003/zaken/api/v1/resultaten/111-111-111"},
        )

        with (
            patch(
                "openarchiefbeheer.destruction.utils.ArchiveConfig.get_solo",
                return_value=ArchiveConfig(**TEST_DATA_ARCHIVE_CONFIG_PARTIAL),
            ),
            patch("openarchiefbeheer.destruction.utils.create_eio_destruction_report"),
            patch("openarchiefbeheer.destruction.utils.attach_report_to_zaak"),
        ):
            destruction_list.create_report_zaak()

        destruction_list.refresh_from_db()

        self.assertEqual(
            destruction_list.zaak_destruction_report_url,
            "http://localhost:8003/zaken/api/v1/zaken/111-111-111",
        )
        self.assertEqual(
            destruction_list.internal_results["created_resources"],
            {
                "resultaten": [
                    "http://localhost:8003/zaken/api/v1/resultaten/111-111-111"
                ]
            },
        )

    @Mocker()
    def test_failure_result_creation(self, m):
        destruction_list = DestructionListFactory.create()
        ServiceFactory.create(
            api_type=APITypes.zrc,
            api_root="http://localhost:8003/zaken/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.ztc,
            api_root="http://localhost:8003/catalogi/api/v1",
        )

        m.get(
            "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
            json={
                "url": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
                "selectielijstklasse": "https://selectielijst.openzaak.nl/api/v1/resultaten/e939c1ad-32e4-409b-a716-6d7d6e7df892",
            },
        )
        m.get(
            "http://localhost:8003/catalogi/api/v1/zaaktypen?identificatie=ZAAKTYPE-2018-0000000002",
            json={
                "count": 1,
                "results": [
                    {
                        "url": "http://localhost:8003/catalogi/api/v1/zaaktypen/ce9feadd-00cb-46c8-a0ef-1d1dfc78586a",
                        "identificatie": "ZAAKTYPE-2018-0000000002",
                        "omschrijving": "Destruction confirmation type",
                        "beginGeldigheid": "2025-03-21",
                    },
                ],
            },
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/zaken",
            json={"url": "http://localhost:8003/zaken/api/v1/zaken/111-111-111"},
        )
        m.post("http://localhost:8003/zaken/api/v1/resultaten", status_code=500)

        with (
            patch(
                "openarchiefbeheer.destruction.utils.ArchiveConfig.get_solo",
                return_value=ArchiveConfig(**TEST_DATA_ARCHIVE_CONFIG),
            ),
            self.assertRaises(HTTPError),
        ):
            destruction_list.create_report_zaak()

        destruction_list.refresh_from_db()

        self.assertEqual(
            destruction_list.zaak_destruction_report_url,
            "http://localhost:8003/zaken/api/v1/zaken/111-111-111",
        )

    @Mocker()
    def test_failure_status_creation(self, m):
        destruction_list = DestructionListFactory.create()
        ServiceFactory.create(
            api_type=APITypes.zrc,
            api_root="http://localhost:8003/zaken/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.ztc,
            api_root="http://localhost:8003/catalogi/api/v1",
        )

        m.get(
            "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
            json={
                "url": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
                "selectielijstklasse": "https://selectielijst.openzaak.nl/api/v1/resultaten/e939c1ad-32e4-409b-a716-6d7d6e7df892",
            },
        )
        m.get(
            "http://localhost:8003/catalogi/api/v1/zaaktypen?identificatie=ZAAKTYPE-2018-0000000002",
            json={
                "count": 1,
                "results": [
                    {
                        "url": "http://localhost:8003/catalogi/api/v1/zaaktypen/ce9feadd-00cb-46c8-a0ef-1d1dfc78586a",
                        "identificatie": "ZAAKTYPE-2018-0000000002",
                        "omschrijving": "Destruction confirmation type",
                        "beginGeldigheid": "2025-03-21",
                    },
                ],
            },
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/zaken",
            json={"url": "http://localhost:8003/zaken/api/v1/zaken/111-111-111"},
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/resultaten",
            json={"url": "http://localhost:8003/zaken/api/v1/resultaten/111-111-111"},
        )
        m.post("http://localhost:8003/zaken/api/v1/statussen", status_code=500)

        with (
            patch(
                "openarchiefbeheer.destruction.utils.ArchiveConfig.get_solo",
                return_value=ArchiveConfig(**TEST_DATA_ARCHIVE_CONFIG),
            ),
            self.assertRaises(HTTPError),
        ):
            destruction_list.create_report_zaak()

        destruction_list.refresh_from_db()

        self.assertEqual(
            destruction_list.zaak_destruction_report_url,
            "http://localhost:8003/zaken/api/v1/zaken/111-111-111",
        )
        self.assertEqual(
            destruction_list.internal_results["created_resources"]["resultaten"],
            ["http://localhost:8003/zaken/api/v1/resultaten/111-111-111"],
        )

    @Mocker()
    def test_failure_document_creation(self, m):
        destruction_list = DestructionListFactory.create(with_report=True)
        ServiceFactory.create(
            api_type=APITypes.ztc,
            api_root="http://localhost:8003/catalogi/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.zrc,
            api_root="http://localhost:8003/zaken/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.drc,
            api_root="http://localhost:8003/documenten/api/v1",
        )
        m.get(
            "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
            json={
                "url": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
                "selectielijstklasse": "https://selectielijst.openzaak.nl/api/v1/resultaten/e939c1ad-32e4-409b-a716-6d7d6e7df892",
            },
        )
        m.get(
            "http://localhost:8003/catalogi/api/v1/zaaktypen?identificatie=ZAAKTYPE-2018-0000000002",
            json={
                "count": 1,
                "results": [
                    {
                        "url": "http://localhost:8003/catalogi/api/v1/zaaktypen/ce9feadd-00cb-46c8-a0ef-1d1dfc78586a",
                        "identificatie": "ZAAKTYPE-2018-0000000002",
                        "omschrijving": "Destruction confirmation type",
                        "beginGeldigheid": "2025-03-21",
                    },
                ],
            },
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/zaken",
            json={"url": "http://localhost:8003/zaken/api/v1/zaken/111-111-111"},
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/resultaten",
            json={"url": "http://localhost:8003/zaken/api/v1/resultaten/111-111-111"},
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/statussen",
            json={"url": "http://localhost:8003/zaken/api/v1/statussen/111-111-111"},
        )
        m.post(
            "http://localhost:8003/documenten/api/v1/enkelvoudiginformatieobjecten",
            status_code=500,
        )

        with (
            patch(
                "openarchiefbeheer.destruction.utils.ArchiveConfig.get_solo",
                return_value=ArchiveConfig(**TEST_DATA_ARCHIVE_CONFIG),
            ),
            self.assertRaises(HTTPError),
        ):
            destruction_list.create_report_zaak()

        destruction_list.refresh_from_db()

        self.assertEqual(
            destruction_list.zaak_destruction_report_url,
            "http://localhost:8003/zaken/api/v1/zaken/111-111-111",
        )
        self.assertEqual(
            destruction_list.internal_results["created_resources"]["resultaten"],
            ["http://localhost:8003/zaken/api/v1/resultaten/111-111-111"],
        )
        self.assertEqual(
            destruction_list.internal_results["created_resources"]["statussen"],
            ["http://localhost:8003/zaken/api/v1/statussen/111-111-111"],
        )

    @Mocker()
    def test_failure_zio_creation(self, m):
        destruction_list = DestructionListFactory.create(with_report=True)
        ServiceFactory.create(
            api_type=APITypes.zrc,
            api_root="http://localhost:8003/zaken/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.drc,
            api_root="http://localhost:8003/documenten/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.ztc,
            api_root="http://localhost:8003/catalogi/api/v1",
        )

        m.get(
            "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
            json={
                "url": "http://localhost:8003/catalogi/api/v1/resultaattypen/5d39b8ac-437a-475c-9a76-0f6ae1540d0e",
                "selectielijstklasse": "https://selectielijst.openzaak.nl/api/v1/resultaten/e939c1ad-32e4-409b-a716-6d7d6e7df892",
            },
        )
        m.get(
            "http://localhost:8003/catalogi/api/v1/zaaktypen?identificatie=ZAAKTYPE-2018-0000000002",
            json={
                "count": 1,
                "results": [
                    {
                        "url": "http://localhost:8003/catalogi/api/v1/zaaktypen/ce9feadd-00cb-46c8-a0ef-1d1dfc78586a",
                        "identificatie": "ZAAKTYPE-2018-0000000002",
                        "omschrijving": "Destruction confirmation type",
                        "beginGeldigheid": "2025-03-21",
                    },
                ],
            },
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/zaken",
            json={"url": "http://localhost:8003/zaken/api/v1/zaken/111-111-111"},
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/resultaten",
            json={"url": "http://localhost:8003/zaken/api/v1/resultaten/111-111-111"},
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/statussen",
            json={"url": "http://localhost:8003/zaken/api/v1/statussen/111-111-111"},
        )
        m.post(
            "http://localhost:8003/documenten/api/v1/enkelvoudiginformatieobjecten",
            json={
                "url": "http://localhost:8003/documenten/api/v1/enkelvoudiginformatieobjecten/111-111-111"
            },
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/zaakinformatieobjecten", status_code=500
        )

        with (
            patch(
                "openarchiefbeheer.destruction.utils.ArchiveConfig.get_solo",
                return_value=ArchiveConfig(**TEST_DATA_ARCHIVE_CONFIG),
            ),
            self.assertRaises(HTTPError),
        ):
            destruction_list.create_report_zaak()

        destruction_list.refresh_from_db()

        self.assertEqual(
            destruction_list.zaak_destruction_report_url,
            "http://localhost:8003/zaken/api/v1/zaken/111-111-111",
        )
        self.assertEqual(
            destruction_list.internal_results["created_resources"]["resultaten"],
            ["http://localhost:8003/zaken/api/v1/resultaten/111-111-111"],
        )
        self.assertEqual(
            destruction_list.internal_results["created_resources"]["statussen"],
            ["http://localhost:8003/zaken/api/v1/statussen/111-111-111"],
        )
        self.assertEqual(
            destruction_list.internal_results["created_resources"][
                "enkelvoudiginformatieobjecten"
            ],
            [
                "http://localhost:8003/documenten/api/v1/enkelvoudiginformatieobjecten/111-111-111"
            ],
        )

    @Mocker()
    def test_resume_after_failure(self, m):
        destruction_list = DestructionListFactory.create(
            zaak_destruction_report_url="http://localhost:8003/zaken/api/v1/zaken/111-111-111"
        )
        result_store = ResultStore(destruction_list)
        internal_results = result_store.get_internal_results()
        # Fake intermediate results that are created during failures
        internal_results["created_resources"].update(
            {
                "resultaten": [
                    "http://localhost:8003/zaken/api/v1/resultaten/111-111-111"
                ],
                "statussen": [
                    "http://localhost:8003/zaken/api/v1/statussen/111-111-111"
                ],
                "enkelvoudiginformatieobjecten": [
                    "http://localhost:8003/documenten/api/v1/enkelvoudiginformatieobjecten/111-111-111"
                ],
            }
        )
        result_store.save()
        ServiceFactory.create(
            api_type=APITypes.zrc,
            api_root="http://localhost:8003/zaken/api/v1",
        )
        ServiceFactory.create(
            api_type=APITypes.ztc,
            api_root="http://localhost:8003/catalogi/api/v1",
        )

        m.get(
            "http://localhost:8003/catalogi/api/v1/zaaktypen?identificatie=ZAAKTYPE-2018-0000000002",
            json={
                "count": 1,
                "results": [
                    {
                        "url": "http://localhost:8003/catalogi/api/v1/zaaktypen/ce9feadd-00cb-46c8-a0ef-1d1dfc78586a",
                        "identificatie": "ZAAKTYPE-2018-0000000002",
                        "omschrijving": "Destruction confirmation type",
                        "beginGeldigheid": "2025-03-21",
                    },
                ],
            },
        )
        m.post(
            "http://localhost:8003/zaken/api/v1/zaakinformatieobjecten",
            json={
                "url": "http://localhost:8003/zaken/api/v1/zaakinformatieobjecten/111-111-111"
            },
        )

        # Calling create_report_zaak resumes from where there is no stored intermediate result (i.e. creating the ZIO)
        with patch(
            "openarchiefbeheer.destruction.utils.ArchiveConfig.get_solo",
            return_value=ArchiveConfig(**TEST_DATA_ARCHIVE_CONFIG),
        ):
            destruction_list.create_report_zaak()

        destruction_list.refresh_from_db()

        self.assertEqual(
            destruction_list.internal_results["created_resources"][
                "zaakinformatieobjecten"
            ],
            ["http://localhost:8003/zaken/api/v1/zaakinformatieobjecten/111-111-111"],
        )

    def test_clean_local_metadata(self):
        destruction_list = DestructionListFactory.create(
            processing_status=InternalStatus.succeeded,
            status=ListStatus.deleted,
            destruction_report=ContentFile(
                b"Hello I am a report.", name="report_test.txt"
            ),
        )

        path = destruction_list.destruction_report.path
        self.assertIsNotNone(destruction_list.destruction_report)
        self.assertTrue(os.path.isfile(path))

        item1 = DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            destruction_list=destruction_list,
            extra_zaak_data={
                "url": "http://zaken.nl/api/v1/zaken/111-111-111",
                "omschrijving": "Test description 1",
                "identificatie": "ZAAK-01",
                "startdatum": "2020-01-01",
                "einddatum": "2022-01-01",
                "resultaat": "http://zaken.nl/api/v1/resultaten/111-111-111",
                "zaaktype": {
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-111",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "nummer": 1,
                    },
                },
            },
        )
        item2 = DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            destruction_list=destruction_list,
            extra_zaak_data={
                "url": "http://zaken.nl/api/v1/zaken/222-222-222",
                "omschrijving": "Test description 2",
                "identificatie": "ZAAK-02",
                "startdatum": "2020-01-01",
                "einddatum": "2022-01-01",
                "resultaat": "http://zaken.nl/api/v1/resultaten/111-111-111",
                "zaaktype": {
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-111",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "nummer": 1,
                    },
                },
            },
        )

        destruction_list.clear_local_metadata()

        item1.refresh_from_db()
        item2.refresh_from_db()

        self.assertEqual(item1.extra_zaak_data, {})
        self.assertEqual(item2.extra_zaak_data, {})
        with self.assertRaises(ValueError):
            destruction_list.destruction_report.file
        self.assertFalse(os.path.isfile(path))


class DestructionListCoReviewTest(TestCase):
    def test_destruction_list_hierarchy(self):
        co_review = DestructionListCoReviewFactory.create()
        self.assertTrue(co_review.created)

        destruction_list = co_review.destruction_list
        self.assertIn(co_review, destruction_list.co_reviews.all())

    def test_author_list_hierarchy(self):
        co_review = DestructionListCoReviewFactory.create()
        author = co_review.author
        self.assertIn(co_review, author.created_co_reviews.all())

    def test_str(self):
        destruction_list = DestructionListFactory(name="Destruction list to co-review")
        author = UserFactory.create(
            username="co-reviewer", first_name="Co", last_name="Reviewer"
        )
        co_review = DestructionListCoReviewFactory.create(
            destruction_list=destruction_list, author=author
        )

        self.assertEqual(
            str(co_review),
            "Co-review for Destruction list to co-review (Co Reviewer (co-reviewer))",
        )

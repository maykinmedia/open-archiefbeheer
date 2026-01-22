from datetime import datetime

from django.contrib.auth.models import Group
from django.test import TestCase
from django.utils import timezone
from django.utils.translation import gettext

from freezegun import freeze_time
from openpyxl import load_workbook
from privates.test import temp_private_root

from openarchiefbeheer.destruction.destruction_report import generate_destruction_report
from openarchiefbeheer.destruction.models import ResourceDestructionResult
from openarchiefbeheer.logging import logevent

from ...accounts.tests.factories import UserFactory
from ..constants import (
    InternalStatus,
    ListItemStatus,
    ListStatus,
    ResourceDestructionResultStatus,
    ReviewDecisionChoices,
)
from .factories import (
    DestructionListFactory,
    DestructionListItemFactory,
    DestructionListReviewFactory,
)


@temp_private_root()
class DestructionReportTests(TestCase):
    def test_generate_destruction_report(self):
        record_manager = UserFactory.create(
            first_name="John",
            last_name="Doe",
            username="jdoe1",
            post__can_start_destruction=True,
        )
        destruction_list = DestructionListFactory.create(
            status=ListStatus.deleted,
            end=datetime(2024, 12, 2, 12, tzinfo=timezone.get_default_timezone()),
        )
        with freeze_time("2024-12-01T12:00:00+01:00"):
            logevent.destruction_list_deletion_triggered(
                destruction_list, record_manager
            )
        item1 = DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            status=ListItemStatus.suggested,
            destruction_list=destruction_list,
        )
        ResourceDestructionResult.objects.create(
            item=item1,
            resource_type="zaken",
            url="http://zaken.nl/api/v1/zaken/111-111-111",
            status=ResourceDestructionResultStatus.deleted,
            metadata={
                "bronapplicatie": "Open Zaak",
                "url": "http://zaken.nl/api/v1/zaken/111-111-111",
                "omschrijving": "Test description 1",
                "identificatie": "ZAAK-01",
                "startdatum": "2020-01-01",
                "einddatum": "2022-01-01",
                "resultaat": {
                    "url": "http://zaken.nl/api/v1/resultaten/111-111-111",
                    "resultaattype": {
                        "url": "http://catalogue-api.nl/catalogi/api/v1/resultaattypen/111-111-111",
                        "archiefactietermijn": "P1D",
                        "omschrijving": "Resulttype 0",
                    },
                },
                "zaaktype": {
                    "uuid": "111-111-111",
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-111",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "naam": "Plannen opstellen",
                        "nummer": 1.1,
                        "jaar": 2020,
                    },
                },
                "selectielijstklasse": "3.2 - Niet vastgesteld - vernietigen",
                "selectielijstklasse_versie": "2020",
            },
        )
        ResourceDestructionResult.objects.create(
            item=item1,
            resource_type="onderwerpobjecten",
            status=ResourceDestructionResultStatus.deleted,
            url="http://localhost:8005/klantinteracties/api/v1/onderwerpobjecten/4b78537d-21f2-4f18-903e-2d3e9114fc7f",
        )
        ResourceDestructionResult.objects.create(
            item=item1,
            resource_type="other_related_resource",
            status=ResourceDestructionResultStatus.unlinked,
            url="http://register.nl/cant/extract/uuid",
        )
        item2 = DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            status=ListItemStatus.suggested,
            destruction_list=destruction_list,
        )
        ResourceDestructionResult.objects.create(
            item=item2,
            resource_type="zaken",
            url="http://zaken.nl/api/v1/zaken/111-111-222",
            status=ResourceDestructionResultStatus.deleted,
            metadata={
                "bronapplicatie": "Open Zaak",
                "url": "http://zaken.nl/api/v1/zaken/111-111-222",
                "omschrijving": "Test description 2",
                "identificatie": "ZAAK-02",
                "startdatum": "2020-01-02",
                "einddatum": "2022-01-02",
                "resultaat": {
                    "url": "http://zaken.nl/api/v1/resultaten/111-111-222",
                    "resultaattype": {
                        "url": "http://catalogue-api.nl/catalogi/api/v1/resultaattypen/111-111-111",
                        "archiefactietermijn": "P1D",
                        "omschrijving": "Resulttype 0",
                    },
                },
                "zaaktype": {
                    "uuid": "111-111-111",
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-111",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "nummer": 1,
                        "naam": "Beleid en regelgeving opstellen",
                        "jaar": 2024,
                    },
                },
                "selectielijstklasse": "3.2 - Niet vastgesteld - vernietigen",
                "selectielijstklasse_versie": "2024",
            },
        )
        item3 = DestructionListItemFactory.create(
            processing_status=InternalStatus.succeeded,
            destruction_list=destruction_list,
        )
        ResourceDestructionResult.objects.create(
            item=item3,
            resource_type="zaken",
            url="http://zaken.nl/api/v1/zaken/111-111-333",
            status=ResourceDestructionResultStatus.deleted,
            metadata={
                "bronapplicatie": "Open Zaak",
                "url": "http://zaken.nl/api/v1/zaken/111-111-333",
                "omschrijving": "Test description 3",
                "identificatie": "ZAAK-03",
                "startdatum": "2020-01-03",
                "einddatum": "2022-01-03",
                "resultaat": None,
                "zaaktype": {
                    "uuid": "111-111-222",
                    "url": "http://catalogi.nl/api/v1/zaaktypen/111-111-222",
                    "omschrijving": "Tralala zaaktype",
                    "selectielijst_procestype": {
                        "naam": "Instellen en inrichten organisatie",
                        "nummer": 1.0,
                        "jaar": 2022,
                    },
                },
                "selectielijstklasse": "1.1 - Ingericht - vernietigen",
                "selectielijstklasse_versie": "2022",
            },
        )

        generate_destruction_report(destruction_list)

        destruction_list.refresh_from_db()

        wb = load_workbook(filename=destruction_list.destruction_report.path)
        sheet_deleted_zaken = wb[gettext("Deleted zaken")]
        rows = list(sheet_deleted_zaken.iter_rows(values_only=True))

        self.assertEqual(len(rows), 4)
        self.assertEqual(
            rows[0],
            (
                "Bronapplicatie",
                "Zaaktype Omschrijving",
                "Zaaktype UUID",
                "Zaaktype Identificatie",
                "Resultaat",
                "Zaak Identificatie",
                "Zaak Startdatum",
                "Zaak Einddatum",
                "Selectielijstklasse",
                "Selectielijst versie",
            ),
        )

        # We don't know the order in which the zaken are retrieved in the db (they are not sorted).
        data_rows = sorted(rows[1:], key=lambda row: row[0])

        self.assertEqual(
            data_rows[0],
            (
                "Open Zaak",
                "Tralala zaaktype",
                "111-111-111",
                None,  # pyopenxl reads empty values as None
                "Resulttype 0",
                "ZAAK-01",
                "2020-01-01",
                "2022-01-01",
                "3.2 - Niet vastgesteld - vernietigen",
                "2020",
            ),
        )
        self.assertEqual(
            data_rows[1],
            (
                "Open Zaak",
                "Tralala zaaktype",
                "111-111-111",
                None,
                "Resulttype 0",
                "ZAAK-02",
                "2020-01-02",
                "2022-01-02",
                "3.2 - Niet vastgesteld - vernietigen",
                "2024",
            ),
        )
        self.assertEqual(
            data_rows[2],
            (
                "Open Zaak",
                "Tralala zaaktype",
                "111-111-222",
                None,
                None,
                "ZAAK-03",
                "2020-01-03",
                "2022-01-03",
                "1.1 - Ingericht - vernietigen",
                "2022",
            ),
        )

        sheet_process_details = wb[gettext("Process details")]
        rows = list(sheet_process_details.iter_rows(values_only=True))

        self.assertEqual(
            rows[0][:5],
            (
                gettext("Date/Time starting destruction"),
                gettext("Date/Time of destruction"),
                gettext("User who started the destruction"),
                gettext("Groups"),
                gettext("Number of deleted cases"),
            ),
        )
        self.assertEqual(
            rows[1][:5],
            (
                "2024-12-01 12:00+01:00",
                "2024-12-02 12:00+01:00",
                "John Doe (jdoe1)",
                None,
                "3",
            ),
        )

        sheet_related_resources = wb[gettext("Related resources")]
        rows = list(sheet_related_resources.iter_rows(values_only=True))

        self.assertEqual(
            rows[0][:3],
            (
                gettext("Resource Type"),
                gettext("Resource UUID"),
                gettext("Operation"),
            ),
        )
        self.assertEqual(
            rows[1][:3],
            (
                "onderwerpobjecten",
                "4b78537d-21f2-4f18-903e-2d3e9114fc7f",
                gettext("Deleted"),
            ),
        )
        self.assertEqual(
            rows[2][:3],
            (
                "other_related_resource",
                "http://register.nl/cant/extract/uuid",
                gettext("Unlinked"),
            ),
        )

    def test_generate_destruction_report_with_multiple_logs(self):
        record_manager1 = UserFactory.create(
            first_name="John",
            last_name="Doe",
            username="jdoe1",
            post__can_start_destruction=True,
        )
        record_manager2 = UserFactory.create(
            first_name="Jane",
            last_name="Doe",
            username="jdoe2",
            post__can_start_destruction=True,
        )
        destruction_list = DestructionListFactory.create(
            status=ListStatus.deleted,
            end=datetime(2024, 12, 2, 12, tzinfo=timezone.get_default_timezone()),
        )
        with freeze_time("2024-12-01T12:00:00+01:00"):
            logevent.destruction_list_deletion_triggered(
                destruction_list, record_manager1
            )
        with freeze_time("2024-12-01T12:30:00+01:00"):
            logevent.destruction_list_deletion_triggered(
                destruction_list, record_manager2
            )

        generate_destruction_report(destruction_list)

        destruction_list.refresh_from_db()

        wb = load_workbook(filename=destruction_list.destruction_report.path)
        sheet_deleted_zaken = wb[gettext("Process details")]
        rows = list(sheet_deleted_zaken.iter_rows(values_only=True))

        self.assertEqual(len(rows), 4)  # No reviews logs
        self.assertEqual(
            rows[0][:5],
            (
                gettext("Date/Time starting destruction"),
                gettext("Date/Time of destruction"),
                gettext("User who started the destruction"),
                gettext("Groups"),
                gettext("Number of deleted cases"),
            ),
        )
        self.assertEqual(
            rows[1][:5],
            (
                "2024-12-01 12:30+01:00",
                "2024-12-02 12:00+01:00",
                "Jane Doe (jdoe2)",
                None,
                "0",
            ),
        )

    def test_generate_destruction_report_review_process(self):
        author = UserFactory.create(post__can_start_destruction=True)
        reviewer = UserFactory.create(
            first_name="John",
            last_name="Doe",
            username="jdoe1",
            post__can_review_destruction=True,
        )
        reviewer_group, created = Group.objects.get_or_create(name="Reviewer")
        reviewer.groups.add(reviewer_group)
        archivist = UserFactory.create(
            first_name="Alice",
            last_name="Wonderland",
            username="awonderland1",
            post__can_review_final_list=True,
        )
        archivist_group, created = Group.objects.get_or_create(name="Archivist")
        archivist.groups.add(archivist_group)

        destruction_list = DestructionListFactory.create(
            status=ListStatus.deleted,
            author=author,
            end=datetime(2024, 12, 2, 12, tzinfo=timezone.get_default_timezone()),
        )

        review_reviewer_rejected = DestructionListReviewFactory.create(
            author=reviewer, decision=ReviewDecisionChoices.rejected
        )
        review_reviewer_accepted = DestructionListReviewFactory.create(
            author=reviewer, decision=ReviewDecisionChoices.accepted
        )
        review_archivist_accepted = DestructionListReviewFactory.create(
            author=archivist, decision=ReviewDecisionChoices.accepted
        )

        logevent.destruction_list_created(
            destruction_list, author, reviewer
        )  # A log with a different template (should NOT be present in the report)

        logevent.destruction_list_reviewed(
            destruction_list,
            review_reviewer_rejected,
            comment="Terrible",
            user=reviewer,
        )  # A rejection (should NOT be present in the report)
        logevent.destruction_list_deletion_triggered(destruction_list, author)
        with freeze_time("2024-05-02T16:00:00+02:00"):
            logevent.destruction_list_reviewed(
                destruction_list,
                review_reviewer_accepted,
                comment="Nice",
                user=reviewer,
            )
        with freeze_time("2024-05-08T09:45:00+02:00"):
            logevent.destruction_list_reviewed(
                destruction_list,
                review_archivist_accepted,
                comment="Ship it",
                user=archivist,
            )

        generate_destruction_report(destruction_list)

        destruction_list.refresh_from_db()

        wb = load_workbook(filename=destruction_list.destruction_report.path)
        sheet_deleted_zaken = wb[gettext("Process details")]
        rows = list(sheet_deleted_zaken.iter_rows(values_only=True))

        self.assertEqual(len(rows), 6)

        self.assertEqual(
            rows[4][:4],
            (
                "Reviewer",
                "John Doe (jdoe1)",
                "2024-05-02 16:00+02:00",
                gettext("Has approved"),
            ),
        )
        self.assertEqual(
            rows[5][:4],
            (
                "Archivist",
                "Alice Wonderland (awonderland1)",
                "2024-05-08 09:45+02:00",
                gettext("Has approved"),
            ),
        )
